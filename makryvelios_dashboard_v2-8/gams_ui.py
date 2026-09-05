"""Streamlit UI for the visible GAMS-compatible ITA Studio."""
from __future__ import annotations

import io
import json
import re
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from gams_compat import (
    GAMS_STATUS_COLOURS,
    RND2437_MC_PRESETS,
    SYN2_WEIGHT_ROUNDS,
    gams_model_text,
    gams_reproducibility_bundle,
    monte_carlo_gams_compatible,
    prepare_gams_compatible_model,
    preset_definition,
    solve_gams_compatible,
    solve_weight_matrix,
)
from llm_bridge import configured as llm_configured, llm_reply, summarise_ita_for_llm


BASE = Path(__file__).resolve().parent
REFERENCE_GAMS = BASE / "reference_gams" / "vangelis"
REFERENCE_GAMS_DATA = REFERENCE_GAMS / "data"

STATUS_STYLE = {
    "GREEN": ("#DCFCE7", "#14532D"),
    "RED": ("#FEE2E2", "#7F1D1D"),
    "GRAY": ("#F3F4F6", "#111827"),
    "GREY": ("#F3F4F6", "#111827"),
    "SELECTED": ("#DBEAFE", "#1E3A8A"),
    "NOT SELECTED": ("#E2E8F0", "#0F172A"),
}


def _suggest(tokens: tuple[str, ...], columns: list[str], excluded: set[str] | None = None) -> str | None:
    excluded = excluded or set()
    for column in columns:
        low = str(column).lower()
        if column not in excluded and any(token in low for token in tokens):
            return column
    return None




def _read_reference_id_set(path: Path) -> set[str]:
    if not path.exists():
        return set()
    raw = path.read_text(encoding="utf-8", errors="ignore")
    return {token.strip().rstrip(",") for token in re.split(r"[\s,;]+", raw) if token.strip().rstrip(",")}


@st.cache_data(show_spinner=False)
def _load_original_syn2_reference() -> pd.DataFrame:
    """Reconstruct the supplied SYN2 540 GAMS input table by project ID."""
    required = [
        REFERENCE_GAMS_DATA / "budget_syn2.prn",
        REFERENCE_GAMS_DATA / "score_syn2.prn",
        REFERENCE_GAMS_DATA / "sector_syn2.prn",
    ]
    missing = [str(path.name) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError("Missing supplied SYN2 reference input(s): " + ", ".join(missing))

    budget = pd.read_csv(required[0], sep=r"\s+", engine="python", index_col=0)
    budget.index = budget.index.astype(str)
    budget.index.name = "project_id"
    budget.columns = [str(c).strip().upper() for c in budget.columns]

    score = pd.read_csv(required[1], sep=r"\s+", engine="python", index_col=0)
    score.index = score.index.astype(str)
    score.index.name = "project_id"
    score.columns = [f"C{i + 1}" for i in range(score.shape[1])]

    sector = pd.read_csv(required[2], sep=r"\s+", engine="python", header=None, names=["project_id", "sector"], dtype={"project_id": str})
    sector["project_id"] = sector["project_id"].astype(str)

    frame = budget.join(score, how="inner").reset_index().merge(sector, on="project_id", how="left", validate="one_to_one")
    frame["ita_status"] = "UNCLASSIFIED"
    status_files = {
        "GREEN": REFERENCE_GAMS_DATA / "green1_syn2.txt",
        "RED": REFERENCE_GAMS_DATA / "red1_syn2.txt",
        "GRAY": REFERENCE_GAMS_DATA / "gray1.txt",
    }
    for status, path in status_files.items():
        ids = _read_reference_id_set(path)
        if ids:
            frame.loc[frame["project_id"].isin(ids), "ita_status"] = status
    return frame

def _read_id_file(upload) -> list[str]:
    if upload is None:
        return []
    text = upload.getvalue().decode("utf-8", errors="ignore")
    ids = []
    for token in re.split(r"[\s,;]+", text):
        token = token.strip().rstrip(",")
        if token:
            ids.append(token)
    return list(dict.fromkeys(ids))


def _dict_editor(title: str, values: dict[str, float], key: str, id_label: str, value_label: str) -> dict[str, float]:
    st.markdown(f"##### {title}")
    frame = pd.DataFrame({id_label: list(values), value_label: list(values.values())})
    if frame.empty:
        frame = pd.DataFrame({id_label: [], value_label: []})
    edited = st.data_editor(
        frame,
        width="stretch",
        hide_index=True,
        num_rows="dynamic",
        column_config={value_label: st.column_config.NumberColumn(value_label, min_value=0.0, format="%.2f")},
        key=key,
    )
    out = {}
    for _, row in edited.iterrows():
        name = str(row.get(id_label, "")).strip()
        if name and pd.notna(row.get(value_label)):
            out[name] = float(row[value_label])
    return out


def _region_mapping_editor(df: pd.DataFrame, preset: dict, key: str) -> tuple[dict[str, str], dict[str, float]]:
    numeric = list(df.select_dtypes(include=np.number).columns)
    rows = []
    regions = preset.get("regions", [])
    if not regions:
        guessed = [c for c in numeric if any(token in str(c).lower() for token in ("budget", "cost", "fund", "att", "cmk", "ep2", "wmk", "ste"))]
        regions = [str(c) for c in guessed[:13]]

    used: set[str] = set()
    budget_like = [c for c in numeric if any(token in str(c).lower() for token in ("budget", "cost", "fund", "att", "cmk", "ep2", "wmk", "ste"))]
    for region in regions:
        exact = next((c for c in numeric if c not in used and str(c).strip().upper() == str(region).upper()), None)
        fuzzy = exact or next((c for c in numeric if c not in used and str(region).lower() in str(c).lower()), None)
        candidate = fuzzy
        if candidate is None and len(regions) == len(budget_like):
            candidate = next((c for c in budget_like if c not in used), None)
        if candidate is not None:
            used.add(candidate)
        rows.append({
            "Region / GAMS rg": region,
            "Source budget column": candidate or "",
            "Budget cap": float(preset.get("region_caps", {}).get(region, 0.0)),
        })

    frame = pd.DataFrame(rows)
    if frame.empty:
        frame = pd.DataFrame([{"Region / GAMS rg": "REGION1", "Source budget column": "", "Budget cap": 0.0}])
    edited = st.data_editor(
        frame,
        width="stretch",
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "Source budget column": st.column_config.SelectboxColumn("Source budget column", options=[""] + numeric, required=False),
            "Budget cap": st.column_config.NumberColumn("Budget cap", min_value=0.0, format="€ %.0f"),
        },
        key=key,
    )
    mapping, caps = {}, {}
    duplicate_sources: dict[str, list[str]] = {}
    source_to_regions: dict[str, list[str]] = {}
    unmapped = []
    for _, row in edited.iterrows():
        region = str(row.get("Region / GAMS rg", "")).strip()
        source = str(row.get("Source budget column", "")).strip()
        if not region:
            continue
        if not source:
            unmapped.append(region)
            continue
        source_to_regions.setdefault(source, []).append(region)
        mapping[region] = source
        if pd.notna(row.get("Budget cap")) and float(row["Budget cap"]) > 0:
            caps[region] = float(row["Budget cap"])

    duplicate_sources = {src: regs for src, regs in source_to_regions.items() if len(regs) > 1}
    if duplicate_sources:
        details = "; ".join(f"{src} → {', '.join(regs)}" for src, regs in duplicate_sources.items())
        st.error("Each GAMS region must use a different source budget column. Duplicate mapping: " + details)
    if unmapped:
        st.warning("Map a source budget column for: " + ", ".join(unmapped) + ". The model will not run until every required region is mapped.")
    return mapping, caps


def _weights_editor(preset_name: str, criteria: list[str]) -> tuple[str, pd.DataFrame | None, list[float] | None]:
    if preset_name == "Vangelis – SYN2 540" and len(criteria) == 3:
        round_name = st.selectbox("Original GAMS weighting round", list(SYN2_WEIGHT_ROUNDS), key="gams_syn2_round")
        matrix = SYN2_WEIGHT_ROUNDS[round_name]
        weight_frame = pd.DataFrame(matrix, index=criteria, columns=["DM1", "DM2", "DM3"])
        editable = weight_frame.reset_index(names="criterion")
        edited = st.data_editor(
            editable,
            width="stretch",
            hide_index=True,
            disabled=["criterion"],
            column_config={c: st.column_config.NumberColumn(c, min_value=0.0, format="%.4f") for c in ["DM1", "DM2", "DM3"]},
            key="gams_syn2_weights",
        )
        out = edited.set_index("criterion")[["DM1", "DM2", "DM3"]].astype(float)
        st.caption("Columns are the individual-decision-maker optimisations used by the original GAMS loop(crit). The application normalises each column only at solve time if necessary.")
        return round_name, out, None
    default = [1 / len(criteria)] * len(criteria)
    vector = pd.DataFrame({"criterion": criteria, "weight": default})
    edited = st.data_editor(
        vector,
        width="stretch",
        hide_index=True,
        disabled=["criterion"],
        column_config={"weight": st.column_config.NumberColumn("Weight", min_value=0.0, format="%.4f")},
        key="gams_custom_weights",
    )
    return "Custom", None, edited.weight.astype(float).tolist()


def _intervention_weight_editor(preset: dict, criteria: list[str]) -> dict[str, list[float]]:
    values = preset.get("intervention_weights", {})
    if not values or len(criteria) != 3:
        return {}
    frame = pd.DataFrame([
        {"Intervention": key, **{criteria[i]: weights[i] for i in range(len(criteria))}}
        for key, weights in values.items()
    ])
    edited = st.data_editor(
        frame,
        width="stretch",
        hide_index=True,
        disabled=["Intervention"],
        column_config={c: st.column_config.NumberColumn(c, min_value=0.0, format="%.4f") for c in criteria},
        key="gams_intervention_weights",
    )
    return {str(row.Intervention): [float(row[c]) for c in criteria] for _, row in edited.iterrows()}


def _styled_projects(frame: pd.DataFrame):
    status_col = "ita_classification" if "ita_classification" in frame else ("ita_status" if "ita_status" in frame else None)
    decision_col = "decision" if "decision" in frame else None

    def cell_style(value):
        key = str(value).upper()
        bg, fg = STATUS_STYLE.get(key, ("#FFFFFF", "#111827"))
        return f"background-color:{bg};color:{fg};font-weight:700"

    styler = frame.style
    if status_col:
        styler = styler.map(cell_style, subset=[status_col])
    if decision_col:
        styler = styler.map(cell_style, subset=[decision_col])
    return styler


def _coloured_excel(projects: pd.DataFrame, tables: dict[str, pd.DataFrame]) -> bytes:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    all_tables = {"Projects": projects, **tables}
    fill_map = {
        "GREEN": PatternFill("solid", fgColor="DCFCE7"),
        "RED": PatternFill("solid", fgColor="FEE2E2"),
        "GRAY": PatternFill("solid", fgColor="F3F4F6"),
        "GREY": PatternFill("solid", fgColor="F3F4F6"),
        "SELECTED": PatternFill("solid", fgColor="DBEAFE"),
        "NOT SELECTED": PatternFill("solid", fgColor="E2E8F0"),
    }
    font_map = {
        "GREEN": "14532D", "RED": "7F1D1D", "GRAY": "111827", "GREY": "111827",
        "SELECTED": "1E3A8A", "NOT SELECTED": "0F172A",
    }
    for name, frame in all_tables.items():
        ws = wb.create_sheet(title=name[:31])
        for j, column in enumerate(frame.columns, start=1):
            cell = ws.cell(row=1, column=j, value=str(column))
            cell.fill = PatternFill("solid", fgColor="0B1F3A")
            cell.font = Font(color="FFFFFF", bold=True)
        for i, (_, row) in enumerate(frame.iterrows(), start=2):
            for j, column in enumerate(frame.columns, start=1):
                value = row[column]
                if isinstance(value, (np.integer, np.floating)):
                    value = value.item()
                cell = ws.cell(row=i, column=j, value=None if pd.isna(value) else value)
                key = str(value).upper()
                if key in fill_map:
                    cell.fill = fill_map[key]
                    cell.font = Font(color=font_map[key], bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _active_llm_config() -> dict:
    return st.session_state.get("llm_config", {})


def render_gams_studio(df: pd.DataFrame) -> None:
    st.subheader("ITA / GAMS-Compatible Portfolio Optimisation Studio")
    st.markdown(
        '<div class="guide"><b>Canonical logic.</b> Evangelos Makryvelios GAMS-style binary portfolio optimisation: sets → parameters/tables → X(p) → budget/sector/intervention constraints → PORTFSCORE → MIP solve.<br>'
        '<b>Execution.</b> The live application executes the equivalent model with the built-in SciPy/HiGHS solver, so a commercial GAMS licence is not required.<br>'
        '<b>Compatibility.</b> GAMS remains visible, exportable and available as an independent replication route. The application does not relabel a HiGHS run as a GAMS run.</div>',
        unsafe_allow_html=True,
    )
    st.info("This is an additive module. The existing 12A.1 ITA-PB / Hybrid ITA-RW module and every earlier analytical workflow remain unchanged.")

    preset_name = st.selectbox("Model template", ["Vangelis – SYN2 540", "Vangelis – R&D 2437", "Custom"], key="gams_preset")
    preset = preset_definition(preset_name)

    if preset_name == "Vangelis – SYN2 540":
        syn2_mode = st.radio(
            "SYN2 input source",
            ["Original supplied GAMS inputs (recommended for exact replication)", "Current application dataset"],
            horizontal=True,
            key="gams_syn2_input_source",
        )
        if syn2_mode.startswith("Original supplied"):
            try:
                df = _load_original_syn2_reference()
                st.success("Loaded the original supplied SYN2 GAMS inputs: 540 projects, EP2/ATT/CMK/WMK/STE budgets, C1-C3 scores, sectors and GREEN/RED/GRAY sets.")
                unclassified = int((df["ita_status"] == "UNCLASSIFIED").sum())
                if unclassified:
                    st.caption(f"{unclassified} supplied project ID is not present in the GREEN/RED/GRAY text sets and remains UNCLASSIFIED; it is not silently reassigned.")
            except Exception as exc:
                st.error(f"Could not load the embedded supplied SYN2 inputs: {exc}")
                return
    elif preset_name == "Vangelis – R&D 2437":
        st.info("The supplied 2,437-project package contains the GAMS model source but not its complete raw budget/score/sector input tables. Map the active dataset or upload those source tables when available.")

    if preset.get("notes"):
        with st.expander("What this preset preserves from the supplied GAMS files", expanded=True):
            for note in preset["notes"]:
                st.markdown(f"- {note}")

    all_columns = list(df.columns)
    numeric = list(df.select_dtypes(include=np.number).columns)
    if not all_columns or not numeric:
        st.error("The GAMS-compatible Studio requires a project table with numeric scores and budgets.")
        return

    st.markdown("### 1. GAMS set and parameter mapping")
    guessed_id = _suggest(("project_id", "project id", "proposal", "application", "project", "κωδ"), all_columns) or all_columns[0]
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        project_id_col = st.selectbox("p · Project ID", all_columns, index=all_columns.index(guessed_id), key="gams_project_id")
    default_criteria = []
    for code in ["c1", "c2", "c3"]:
        candidate = next((c for c in numeric if str(c).strip().lower() == code), None)
        if candidate:
            default_criteria.append(candidate)
    if len(default_criteria) < 3:
        default_criteria = [c for c in numeric if c != project_id_col][:3]
    with c2:
        criteria = st.multiselect("crit · Criterion columns", numeric, default=default_criteria, max_selections=12, key="gams_criteria")
    sector_guess = _suggest(("sector", "θεμα", "sector_syn"), all_columns)
    with c3:
        sector_options = [None] + all_columns
        sector_col = st.selectbox("sec · Sector (optional)", sector_options, index=sector_options.index(sector_guess) if sector_guess in sector_options else 0, key="gams_sector")
    intervention_guess = _suggest(("intervention", "intv", "παρέμβ"), all_columns)
    with c4:
        intervention_options = [None] + all_columns
        intervention_col = st.selectbox("intv · Intervention (optional)", intervention_options, index=intervention_options.index(intervention_guess) if intervention_guess in intervention_options else 0, key="gams_intervention")

    if not criteria:
        st.warning("Map at least one criterion to continue.")
        return

    region_mapping, region_caps = _region_mapping_editor(df, preset, "gams_region_mapping")
    if not region_mapping:
        st.warning("Map at least one regional/project budget column. The original models formulate expenditure through budget(p,rg).")
        return

    left, middle, right = st.columns(3)
    with left:
        sector_caps = _dict_editor("Sector ceilings · totbudg(sec)", preset.get("sector_caps", {}) if sector_col else {}, "gams_sector_caps", "Sector", "Cap")
    with middle:
        intervention_caps = _dict_editor("Intervention ceilings · totbudgi(intv)", preset.get("intervention_caps", {}) if intervention_col else {}, "gams_intv_caps", "Intervention", "Cap")
    with right:
        group_caps = _dict_editor("Region-group ceilings", preset.get("region_group_caps", {}), "gams_group_caps", "Region group", "Cap")

    st.markdown("### 2. GREEN / RED / GRAY and X.fx logic")
    status_guess = _suggest(("ita_status", "classification", "colour", "color", "green", "status"), all_columns)
    s1, s2, s3, s4 = st.columns(4)
    with s1:
        status_options = [None] + all_columns
        status_col = st.selectbox("Classification/status column", status_options, index=status_options.index(status_guess) if status_guess in status_options else 0, key="gams_status_col")
    with s2:
        fix_green = st.checkbox("X.fx(GREEN) = 1", value=preset_name != "Custom", key="gams_fix_green")
    with s3:
        fix_red = st.checkbox("X.fx(RED) = 0", value=preset_name != "Custom", key="gams_fix_red")
    with s4:
        mip_gap = st.number_input("MIP relative gap / optcr", min_value=0.0, max_value=0.1, value=0.0005 if preset_name == "Vangelis – R&D 2437" else 0.0, step=0.0005, format="%.4f", key="gams_mip_gap")

    u1, u2 = st.columns(2)
    with u1:
        green_file = st.file_uploader("Optional GREEN ID set (.txt/.prn)", type=["txt", "prn"], key="gams_green_file")
    with u2:
        red_file = st.file_uploader("Optional RED ID set (.txt/.prn)", type=["txt", "prn"], key="gams_red_file")
    fixed_in = _read_id_file(green_file)
    fixed_out = _read_id_file(red_file)

    default_factors = {}
    if preset_name == "Vangelis – R&D 2437":
        default_factors = {"GREY": 0.925, "NEWGREEN": 0.925, "GREY2": 0.85}
    factor_frame = pd.DataFrame({"Status set": list(default_factors), "Effective budget factor": list(default_factors.values())})
    with st.expander("Budget-adjustment sets used in later GAMS rounds", expanded=preset_name == "Vangelis – R&D 2437"):
        st.caption("These reproduce rules such as loop(GREY, budget(GREY,rg)=0.925*budget(GREY,rg)). Leave a status absent if that rule is not active in the current round.")
        factor_edited = st.data_editor(
            factor_frame,
            width="stretch",
            hide_index=True,
            num_rows="dynamic",
            column_config={"Effective budget factor": st.column_config.NumberColumn(min_value=0.01, max_value=2.0, format="%.3f")},
            key="gams_budget_factors",
        )
    budget_factors = {str(row["Status set"]).strip().upper(): float(row["Effective budget factor"]) for _, row in factor_edited.iterrows() if str(row.get("Status set", "")).strip() and pd.notna(row.get("Effective budget factor"))}

    st.markdown("### 3. Original weighting logic")
    round_name, weight_matrix, weight_vector = _weights_editor(preset_name, [f"C{i + 1}" for i in range(len(criteria))])
    intervention_weights = _intervention_weight_editor(preset, [f"C{i + 1}" for i in range(len(criteria))]) if intervention_col else {}
    if intervention_weights:
        st.caption("For the 2,437-project preset, these intervention-specific vectors reproduce the original GAMS loop(p) objective coefficients.")

    metadata = {
        "preset": preset_name,
        "source_round": round_name,
        "fix_green": bool(fix_green),
        "fix_red": bool(fix_red),
        "mip_rel_gap": float(mip_gap),
        "canonical_logic": "Evangelos Makryvelios GAMS-compatible portfolio MIP",
    }
    if preset_name == "Vangelis – SYN2 540":
        metadata["source_model_note"] = "CMK differs by €1,000 between supplied first-round and later source files; configured cap is shown explicitly in the UI."

    st.markdown("### 4. Build and solve")
    expected_regions = list(preset.get("regions", []))
    duplicate_budget_columns = len(set(region_mapping.values())) != len(region_mapping)
    missing_regions = [region for region in expected_regions if region not in region_mapping]
    mapping_ready = bool(region_mapping) and not duplicate_budget_columns and not missing_regions
    if duplicate_budget_columns:
        st.error("Resolve duplicate regional budget mappings above before solving.")
    if missing_regions:
        st.warning("Required GAMS region mapping still missing for: " + ", ".join(missing_regions))
    r1, r2 = st.columns([2, 1])
    with r1:
        run_clicked = st.button("RUN GAMS-COMPATIBLE MODEL WITH HiGHS", type="primary", key="gams_run", disabled=not mapping_ready)
    with r2:
        st.caption("The solver is HiGHS. The model formulation and exports remain GAMS-style and auditable.")

    if run_clicked:
        try:
            model = prepare_gams_compatible_model(
                df,
                project_id_column=project_id_col,
                criterion_columns=criteria,
                region_budget_columns=region_mapping,
                region_caps=region_caps,
                sector_column=sector_col,
                sector_caps=sector_caps,
                intervention_column=intervention_col,
                intervention_caps=intervention_caps,
                status_column=status_col,
                fixed_in=fixed_in,
                fixed_out=fixed_out,
                budget_factors=budget_factors,
                region_groups=preset.get("region_groups", {}),
                region_group_caps=group_caps,
                intervention_weights=intervention_weights,
                metadata=metadata,
            )
            if weight_matrix is not None:
                summary, result_map = solve_weight_matrix(model, weight_matrix, mip_rel_gap=float(mip_gap))
                selection_matrix = result_map.pop("__selection_matrix__")
                counts = selection_matrix.drop(columns=["project_id"]).sum(axis=1)
                total_scenarios = selection_matrix.shape[1] - 1
                selection_matrix["ITA classification"] = np.where(counts == total_scenarios, "GREEN", np.where(counts == 0, "RED", "GRAY"))
                st.session_state["gams_model"] = model
                st.session_state["gams_run_summary"] = summary
                st.session_state["gams_result_map"] = result_map
                st.session_state["gams_selection_matrix"] = selection_matrix
                st.session_state["gams_active_scenario"] = list(result_map)[0]
                st.session_state["gams_weight_matrix"] = weight_matrix
                st.session_state["gams_weight_vector"] = None
            else:
                run = solve_gams_compatible(model, weights=weight_vector, mip_rel_gap=float(mip_gap))
                st.session_state["gams_model"] = model
                st.session_state["gams_run_summary"] = pd.DataFrame([{
                    "scenario": round_name, "solver_status": run.status, "portfolio_score": run.objective,
                    "selected_projects": int(run.project_results.selected.sum()), "allocated_budget": float(run.project_results.allocated_budget.sum()),
                }])
                st.session_state["gams_result_map"] = {round_name: run}
                st.session_state["gams_selection_matrix"] = pd.DataFrame()
                st.session_state["gams_active_scenario"] = round_name
                st.session_state["gams_weight_matrix"] = None
                st.session_state["gams_weight_vector"] = weight_vector
            st.success("Model compiled and solved. GAMS logic is preserved; execution completed with SciPy/HiGHS.")
        except Exception as exc:
            st.error(f"GAMS-compatible solve failed: {exc}")

    model = st.session_state.get("gams_model")
    result_map = st.session_state.get("gams_result_map", {})
    if model is None or not result_map:
        st.info("Configure the model and run it to activate outputs, GAMS export, Monte Carlo and LLM interpretation.")
        _source_gams_library()
        return

    scenario_names = list(result_map)
    active_scenario = st.selectbox("Displayed solved scenario", scenario_names, index=scenario_names.index(st.session_state.get("gams_active_scenario")) if st.session_state.get("gams_active_scenario") in scenario_names else 0, key="gams_display_scenario")
    st.session_state["gams_active_scenario"] = active_scenario
    run = result_map[active_scenario]
    active_weights = None
    weight_matrix_saved = st.session_state.get("gams_weight_matrix")
    if isinstance(weight_matrix_saved, pd.DataFrame) and active_scenario in weight_matrix_saved:
        active_weights = weight_matrix_saved[active_scenario].astype(float).tolist()
    else:
        active_weights = st.session_state.get("gams_weight_vector")

    tabs = st.tabs(["Overview", "Project matrix", "Constraints & allocations", "GAMS model", "Monte Carlo", "LLM co-pilot", "Original source GAMS"])
    with tabs[0]:
        summary = st.session_state.get("gams_run_summary", pd.DataFrame())
        st.dataframe(summary, width="stretch", hide_index=True)
        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric("Solver", "HiGHS")
        k2.metric("Status", run.status)
        k3.metric("Selected projects", f"{int(run.project_results.selected.sum()):,}")
        k4.metric("PORTFSCORE", f"{run.objective:,.3f}")
        k5.metric("Allocated budget", f"€{run.project_results.allocated_budget.sum():,.0f}")
        st.caption("GAMS-compatible formulation / HiGHS execution. Numerical selection is computed by the solver; no LLM participates in the optimisation.")
        if not run.region_allocation.empty:
            fig = px.bar(run.region_allocation, x="region", y="allocated_budget", color="utilisation", title="Regional allocation and constraint utilisation", color_continuous_scale="Blues")
            fig.update_layout(height=480)
            st.plotly_chart(fig, width="stretch")
        selection_matrix = st.session_state.get("gams_selection_matrix", pd.DataFrame())
        if not selection_matrix.empty:
            st.markdown("##### Trichotomy across the original individual-decision-maker optimisations")
            st.dataframe(_styled_projects(selection_matrix), width="stretch", hide_index=True)

    with tabs[1]:
        display_cols = [c for c in ["project_id", *model.criteria, "sector", "intervention", "ita_status", "weighted_score", "effective_budget", "selected", "decision"] if c in run.project_results]
        table = run.project_results[display_cols].copy()
        st.dataframe(_styled_projects(table), width="stretch", hide_index=True, height=650)
        st.download_button("Download project decisions (CSV)", run.project_results.to_csv(index=False).encode("utf-8-sig"), "gams_compatible_project_decisions.csv", "text/csv", key="gams_project_csv")
        excel = _coloured_excel(run.project_results, {"Regions": run.region_allocation, "Sectors": run.sector_allocation, "Interventions": run.intervention_allocation, "Constraints": run.constraint_diagnostics})
        st.download_button("Download ITA-coloured Excel", excel, "gams_compatible_ita_results.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="gams_coloured_excel")

    with tabs[2]:
        st.markdown("#### WHAT LIMITED THE PORTFOLIO?")
        if run.constraint_diagnostics.empty:
            st.info("No active budget/sector/intervention ceilings were configured.")
        else:
            diag = run.constraint_diagnostics.sort_values("utilisation", ascending=False)
            st.dataframe(diag, width="stretch", hide_index=True)
            fig = px.bar(diag.head(30).sort_values("utilisation"), x="utilisation", y="constraint", orientation="h", color="binding", title="Most binding constraints")
            fig.update_xaxes(tickformat=".0%")
            st.plotly_chart(fig, width="stretch")
        if not run.region_allocation.empty:
            st.markdown("##### Regions")
            st.dataframe(run.region_allocation, width="stretch", hide_index=True)
        if not run.sector_allocation.empty:
            st.markdown("##### Sectors")
            st.dataframe(run.sector_allocation, width="stretch", hide_index=True)
        if not run.intervention_allocation.empty:
            st.markdown("##### Interventions")
            st.dataframe(run.intervention_allocation, width="stretch", hide_index=True)

    with tabs[3]:
        st.markdown("#### Live GAMS view")
        gams_text = gams_model_text(model, weights=active_weights)
        st.code(gams_text, language="text", line_numbers=True)
        d1, d2 = st.columns(2)
        with d1:
            st.download_button("Download generated model.gms", gams_text.encode("utf-8"), "model.gms", "text/plain", key="gams_model_download")
        with d2:
            bundle = gams_reproducibility_bundle(model, weights=active_weights, run=run)
            st.download_button("Download complete GAMS-compatible package", bundle, "gams_compatible_complete_package.zip", "application/zip", key="gams_bundle_download")
        st.warning("This is a GAMS-compatible export. The live result above was solved by HiGHS, not by a hidden GAMS installation.")

    with tabs[4]:
        _render_monte_carlo(model, run, active_weights, preset_name)

    with tabs[5]:
        _render_llm_copilot(run)

    with tabs[6]:
        _source_gams_library()


def _render_monte_carlo(model, run, active_weights, preset_name: str) -> None:
    st.markdown("#### Robustness / Monte Carlo")
    if preset_name == "Vangelis – R&D 2437":
        mc_name = st.selectbox("Original GAMS Monte Carlo preset", list(RND2437_MC_PRESETS), key="gams_mc_preset")
        defaults = RND2437_MC_PRESETS[mc_name]
    else:
        mc_name = "Custom"
        defaults = {"iterations": 1000, "seed": 5780, "step": 0.5, "integer_low": -2, "integer_high": 2}
    a, b, c, d = st.columns(4)
    with a:
        iterations = st.number_input("Iterations", 1, 5000, int(defaults.get("iterations", 1000)), 1, key="gams_mc_iterations")
    with b:
        seed = st.number_input("Random seed", 0, 2_147_483_647, int(defaults.get("seed", 5780)), 1, key="gams_mc_seed")
    with c:
        step = st.number_input("Score perturbation step", 0.0, 5.0, float(defaults.get("step", 0.5)), 0.1, key="gams_mc_step")
    with d:
        gap = st.number_input("MC MIP gap", 0.0, 0.1, float(run.settings.get("mip_rel_gap", 0.0005)), 0.0005, format="%.4f", key="gams_mc_gap")
    e, f, g, h = st.columns(4)
    with e:
        integer_low = st.number_input("Uniform integer low", -10, 0, int(defaults.get("integer_low", -2)), 1, key="gams_mc_low")
    with f:
        integer_high = st.number_input("Uniform integer high", 0, 10, int(defaults.get("integer_high", 2)), 1, key="gams_mc_high")
    with g:
        green = st.slider("GREEN threshold", 0.50, 1.00, 0.99, 0.01, key="gams_mc_green")
    with h:
        red = st.slider("RED threshold", 0.00, 0.50, 0.01, 0.01, key="gams_mc_red")
    if st.button("RUN MONTE CARLO", key="gams_mc_run"):
        try:
            with st.spinner("Running the configured GAMS-style Monte Carlo through HiGHS..."):
                projects, draws = monte_carlo_gams_compatible(
                    model, weights=active_weights, iterations=int(iterations), seed=int(seed), perturbation_step=float(step),
                    integer_low=int(integer_low), integer_high=int(integer_high), green_threshold=float(green), red_threshold=float(red), mip_rel_gap=float(gap),
                )
            st.session_state["gams_mc_projects"] = projects
            st.session_state["gams_mc_draws"] = draws
            st.session_state["gams_mc_settings"] = {
                "preset": mc_name, "iterations": int(iterations), "seed": int(seed), "step": float(step),
                "integer_low": int(integer_low), "integer_high": int(integer_high), "green_threshold": float(green), "red_threshold": float(red), "mip_rel_gap": float(gap),
            }
            st.success("Monte Carlo complete.")
        except Exception as exc:
            st.error(f"Monte Carlo failed: {exc}")
    projects = st.session_state.get("gams_mc_projects")
    draws = st.session_state.get("gams_mc_draws")
    if isinstance(projects, pd.DataFrame) and isinstance(draws, pd.DataFrame):
        counts = projects.ita_classification.value_counts().rename_axis("classification").reset_index(name="projects")
        fig = px.bar(counts, x="classification", y="projects", color="classification", color_discrete_map={"GREEN": "#16A34A", "RED": "#DC2626", "GRAY": "#6B7280"}, title="Monte Carlo GREEN / GRAY / RED classification")
        st.plotly_chart(fig, width="stretch")
        st.dataframe(_styled_projects(projects.sort_values("selection_frequency", ascending=False)), width="stretch", hide_index=True, height=620)
        hist = px.histogram(draws, x="portfolio_score", nbins=50, title="Portfolio-score distribution across Monte Carlo runs")
        st.plotly_chart(hist, width="stretch")
        settings = st.session_state.get("gams_mc_settings", {})
        bundle = gams_reproducibility_bundle(model, weights=active_weights, run=run, monte_carlo_projects=projects, monte_carlo_draws=draws, monte_carlo_settings=settings)
        st.download_button("Download model + Monte Carlo reproducibility package", bundle, "gams_monte_carlo_reproducibility.zip", "application/zip", key="gams_mc_bundle")


def _render_llm_copilot(run) -> None:
    st.markdown("#### LLM interpretation co-pilot")
    config = _active_llm_config()
    if not llm_configured(config):
        st.info("Enter your LLM API Key in the always-visible sidebar panel. The optimiser remains fully operational without an LLM.")
        return
    st.success(f"LLM connected for this session: {config.get('provider')} · {config.get('model')}")
    st.caption("The LLM does not select projects or alter constraints. When you press the button below, the application sends a compact computed-results summary for interpretation/drafting.")
    task = st.text_area(
        "What should the LLM do with this computed run?",
        value="Interpret the optimisation result, identify the most binding constraints, explain the regional/sector/intervention allocation, distinguish robust computed findings from methodological caveats, and draft a concise publication-ready Results paragraph.",
        key="gams_llm_task",
        height=150,
    )
    if st.button("ASK LLM ABOUT THIS RUN", key="gams_llm_ask"):
        try:
            evidence = summarise_ita_for_llm(run)
            prompt = f"USER TASK:\n{task}\n\nCOMPUTED EVIDENCE (do not change numbers):\n{evidence}"
            answer = llm_reply(prompt, config)
            st.session_state["gams_llm_answer"] = answer
        except Exception as exc:
            st.error(f"LLM request failed: {exc}")
    if st.session_state.get("gams_llm_answer"):
        st.markdown("##### AI INTERPRETATION — not a computed solver output")
        st.markdown(st.session_state["gams_llm_answer"])


def _source_gams_library() -> None:
    st.markdown("#### Original supplied GAMS source library")
    files = sorted(REFERENCE_GAMS.glob("*.gms")) if REFERENCE_GAMS.exists() else []
    if not files:
        st.info("No bundled reference GAMS sources are available in this package.")
        return
    names = [f.name for f in files]
    selected = st.selectbox("Reference source", names, key="gams_source_reference")
    path = REFERENCE_GAMS / selected
    text = path.read_text(encoding="utf-8", errors="replace")
    st.code(text, language="text", line_numbers=True)
    st.download_button("Download original supplied .gms", text.encode("utf-8"), selected, "text/plain", key="gams_source_download")
